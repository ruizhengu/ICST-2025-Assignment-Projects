import json
import os
import re
import shutil
import subprocess
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

INTRO_CLASS_PATH = Path("/Users/ruizhengu/Experiments/APR-as-AAT/IntroClassJava/dataset")
PATCH_RESULTS = Path("/Users/ruizhengu/Projects/APR-as-AAT/repair-generation/results/results_introclass.xlsx")


def get_intro_class_datasets():
    outputs = {}
    datasets = list(INTRO_CLASS_PATH.glob('*'))
    for dataset in datasets:
        if dataset.is_dir():
            outputs[dataset.name] = dataset
    return outputs


def run_cmd(command):
    process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = process.communicate()
    output = stdout.decode('utf-8')
    errors = stderr.decode('utf-8')
    return "\n".join([output, errors])


def replace_package(file, old, new):
    """
    Replacing the package

    :param file: path of the file for replacing
    :param old: the targeted string to be replaced
    :param new: the new string
    :return:
    """
    file_path = Path(file)
    if not is_valid_file(file_path):
        print("Please enter a valid path")
        return
    with file_path.open("r") as f:
        content = f.readlines()
    with file_path.open("w") as f:
        for line in content:
            if old in line:
                # line = f"{new}\n"
                line = line.replace(old, new)
            f.write(line)


def import_class(file, line_before, new_line):
    """
    Importing a new class after a specific class

    :param file: path of the file for replacing
    :param line_before: the line before the inserting
    :param new_line: the new line to import
    :return:
    """
    file_path = Path(file)
    if not is_valid_file(file_path):
        print("Please enter a valid path")
        return
    with file_path.open("r") as f:
        content = f.readlines()
    with file_path.open("w") as f:
        for line in content:
            f.write(line)
            if line.startswith(line_before):
                f.write(f"{new_line}\n")


def create_excel(file):
    """
    Create an empty xlsx file

    :param file: the path of the file
    """
    wb = Workbook()
    if len(wb.sheetnames) > 0:
        wb.remove(wb.active)
    wb.create_sheet("workspace")
    wb.save(file)


def append_excel(file, data):
    """
    Append data in the xlsx file

    :param file: the path of the xlsx file
    :param data: the new data to append
    """
    try:
        df = pd.read_excel(file, sheet_name="workspace")
    except FileNotFoundError:
        df = pd.DataFrame()
    df = pd.concat([df, pd.DataFrame(data, index=[0])])
    df.to_excel(file, index=False, sheet_name="workspace")


def apply_patch(astor_output):
    with astor_output.open("r") as f:
        data = json.load(f)
        path, modified_path = get_max_suspicious(data)
    shutil.move(modified_path, path)
    # class_name = Path(modified_path).name.replace(".java", "")
    # new_content = get_class_content(modified_path, class_name)
    # replace_class(path, class_name, new_content)
    # run_cmd(f"mvn -f {path} compile")
    # run_cmd(f"mvn -f {path} test")


def get_max_suspicious(data):
    max_suspicious = -1
    path_max = ""
    modified_path_max = ""
    for patch in data["patches"]:
        for patch_hunk in patch["patchhunks"]:
            suspicious = float(patch_hunk["SUSPICIOUNESS"])
            if suspicious > max_suspicious:
                max_suspicious = suspicious
                path_max = patch_hunk["PATH"]
                modified_path_max = patch_hunk["MODIFIED_FILE_PATH"]

    return path_max, modified_path_max


def update_patch_paths(folder):
    """
    Update the path in the astor_output.json file to the renamed unique path

    :param folder: path of the output folder
    """
    file_path = Path(folder) / "astor_output.json"
    with file_path.open('r') as f:
        data = json.load(f)
    for patch in data["patches"]:
        for patch_hunk in patch["patchhunks"]:
            path = patch_hunk["PATH"].replace("\\/", "/")
            modified_path = os.path.normpath(patch_hunk["MODIFIED_FILE_PATH"].replace("\\/", "/"))
            path_parts = modified_path.split(os.sep)
            folder_path = str(folder).split(os.sep)
            # modified_path = os.path.join(os.sep.join(folder_path), os.sep.join(path_parts[9:]))
            modified_path = os.path.join(os.sep.join(folder_path), os.sep.join(path_parts[11:]))
            patch_hunk["PATH"] = path
            patch_hunk["MODIFIED_FILE_PATH"] = modified_path
    with file_path.open('w') as f:
        json.dump(data, f, indent=2)


def is_valid_file(file):
    return Path(file).is_file()


def get_class_content(file, class_name):
    with open(file, 'r') as f:
        content = f.read()
    pattern = rf"class\s+{re.escape(class_name)}\s*\{{(.*?)\}}$"
    match = re.search(pattern, content, re.DOTALL)
    if match:
        return match.group(0)
    else:
        print(f"Class '{class_name}' not found in {file}")
        return None


def replace_class(file, class_name, class_content):
    pattern = rf"class\s+{re.escape(class_name)}\s*\{{(.*?)\}}$"
    newline_pattern = r'(?<!\\)\\n'
    with open(file, "r") as f:
        content = f.read()
    class_content = re.sub(newline_pattern, r'\\\\n', class_content)
    modified_content = re.sub(pattern, class_content, content, flags=re.DOTALL)
    with open(file, "w") as f:
        f.write(modified_content)


def empty_directory(path):
    for item in path.iterdir():
        if item.is_dir():
            shutil.rmtree(item)
        else:
            item.unlink()


def gradle_get_tests(submission):
    """
    Get positive test classes, negative test classes and the number of total test classes.
    :param submission:
    :return:
    """
    cmd = f"{submission}/gradlew clean test -p {submission}"
    output = run_cmd(cmd)
    pattern_positive = re.compile(r"\n(\w+) > .+ PASSED")
    pattern_negative = re.compile(r"\n(\w+) > .+ FAILED")
    pattern_all = re.compile(r"\n(\w+) > .+ (?:FAILED|PASSED)")
    positive_tests = set(pattern_positive.findall(output))
    negative_tests = set(pattern_negative.findall(output))
    num_tests = len(set(pattern_all.findall(output)))
    return positive_tests, negative_tests, num_tests
