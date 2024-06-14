import re
import shutil
import subprocess

import pandas as pd
from openpyxl import Workbook


def run_cmd(command):
    process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = process.communicate()
    output = stdout.decode('utf-8')
    errors = stderr.decode('utf-8')
    return "\n".join([output, errors])


def create_excel(file):
    """
    Create an empty xlsx file

    :param file: the path of the file
    """
    wb = Workbook()
    if len(wb.sheetnames) > 0:
        wb.remove(wb.active)
    wb.create_sheet("workspace")
    wb.save(file)


def append_excel(file, data):
    """
    Append data in the xlsx file

    :param file: the path of the xlsx file
    :param data: the new data to append
    """
    try:
        df = pd.read_excel(file, sheet_name="workspace")
    except FileNotFoundError:
        df = pd.DataFrame()
    df = pd.concat([df, pd.DataFrame(data, index=[0])])
    df.to_excel(file, index=False, sheet_name="workspace")


def empty_directory(path, keep=None):
    for item in path.iterdir():
        if keep is not None and item.name == keep:
            continue
        if item.is_dir():
            shutil.rmtree(item)
        else:
            item.unlink()


def gradle_get_tests(submission):
    """
    Get positive test classes, negative test classes and the number of total test classes.
    :param submission:
    :return:
    """
    cmd = f"{submission}/gradlew clean test -p {submission}"
    output = run_cmd(cmd)
    pattern_positive = re.compile(r"\n(\w+) > .+ PASSED")
    pattern_negative = re.compile(r"\n(\w+) > .+ FAILED")
    pattern_all = re.compile(r"\n(\w+) > .+ (?:FAILED|PASSED)")
    positive_tests = set(pattern_positive.findall(output))
    negative_tests = set(pattern_negative.findall(output))
    num_tests = len(set(pattern_all.findall(output)))
    return positive_tests, negative_tests, num_tests
